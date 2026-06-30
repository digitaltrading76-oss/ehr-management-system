from openpyxl import load_workbook
import io

COMPANY_POLICY_RULES = [
    {"keywords":["late","tardy","tardiness"],"policy_category":"Timekeeping and Attendance","possible_violation":"Tardiness","policy_basis":"Company Code of Conduct - Tardiness.","usual_next_step":"Verify attendance logs and prior offense count."},
    {"keywords":["awol","absent","unauthorized absence"],"policy_category":"Timekeeping and Attendance","possible_violation":"AWOL / Unauthorized Absence","policy_basis":"Company Code of Conduct - AWOL and unauthorized absences.","usual_next_step":"Verify schedule, leave filing, and notice to supervisor/HR."},
    {"keywords":["on hold","parcel","gps","timestamp","delivery","damage","wrong tagging"],"policy_category":"Performance of Duties","possible_violation":"Possible Simple Negligence / Performance Issue","policy_basis":"Company Code of Conduct - Negligence in the Performance of Duties.","usual_next_step":"Verify delivery app logs, GPS/timestamp records, route history, and worker explanation."},
    {"keywords":["theft","steal","stolen","fraud","falsified","cash","remittance","missing parcel"],"policy_category":"Behavior / Company Property","possible_violation":"Possible Theft, Fraud, Dishonesty, or Breach of Trust","policy_basis":"Company Code of Conduct - stealing, falsification, breach of trust and confidence, or company property violations.","usual_next_step":"Secure evidence chain, inventory/cash logs, witness statements, and employee explanation."},
    {"keywords":["accident","injury","hospital","medical","crash"],"policy_category":"Health, Safety and Security","possible_violation":"Safety / Accident Incident","policy_basis":"Company Code of Conduct - Health, Safety and Security obligations.","usual_next_step":"Prioritize safety documentation, medical report, accident report, and welfare verification."},
    {"keywords":["salary","wage","deduction","underpaid","payroll","unpaid"],"policy_category":"Labor Standards / Payroll Concern","possible_violation":"Salary or Wage Complaint","policy_basis":"Requires payroll, attendance, and labor standards verification.","usual_next_step":"Validate payroll computation, attendance, deductions, and payout proof."}
]

LABOR_STANDARDS = [
    {"keywords":["theft","fraud","falsified","dishonesty","cash","remittance"],"labor_category":"Possible serious misconduct / fraud / breach of trust","legal_note":"May fall under just cause principles only if supported by substantial evidence and due process."},
    {"keywords":["negligence","on hold","parcel","gps","timestamp","damage"],"labor_category":"Possible neglect of duties","legal_note":"Neglect must be supported by facts and proportional to the proposed action."},
    {"keywords":["salary","wage","deduction","underpaid","payroll","unpaid"],"labor_category":"Labor standards/payroll compliance concern","legal_note":"Requires payroll verification and may not be a worker violation."},
    {"keywords":["accident","injury","hospital","medical","crash"],"labor_category":"Workplace safety / welfare concern","legal_note":"Focus first on documentation, assistance, and safety compliance."}
]

def extract_text_from_file(filename, content):
    name=(filename or "").lower()
    try:
        if name.endswith((".txt",".csv")):
            return content.decode("utf-8", errors="ignore")
        if name.endswith(".xlsx"):
            wb=load_workbook(io.BytesIO(content), data_only=True)
            out=[]
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    out.append(" ".join("" if v is None else str(v) for v in row))
            return "\n".join(out)
        if name.endswith((".jpg",".jpeg",".png")):
            return f"[Image uploaded: {filename}. OCR/handwriting extraction will be added in production.]"
        if name.endswith((".pdf",".doc",".docx")):
            return f"[Document uploaded: {filename}. Full extraction will be added in production.]"
    except Exception as e:
        return f"[Could not extract text from {filename}: {e}]"
    return f"[Uploaded file: {filename}]"

def has_any(text, words):
    return any(w in text for w in words)

def assess_case(worker_name, worker_type, category, extracted_text, uploaded_names, submitted_by, area):
    text=f"{category} {extracted_text} {' '.join(uploaded_names)}".lower()
    policy=[r for r in COMPANY_POLICY_RULES if has_any(text,r["keywords"])]
    labor=[r for r in LABOR_STANDARDS if has_any(text,r["keywords"])]

    if not policy:
        policy=[{"policy_category":"No exact company policy match","possible_violation":"Unclassified or policy-silent incident","policy_basis":"No direct match found in coded policy rules. Central Command should manually review the company policy and validate under labor standards.","usual_next_step":"Request more facts and supporting documents before formal action."}]
    if not labor:
        labor=[{"labor_category":"General labor due process review","legal_note":"No specific legal category identified. Apply due process and request facts before any action."}]

    names=" ".join(uploaded_names).lower()
    checklist=[
        {"item":"Incident report uploaded","status":has_any(names,["incident","ir","report"]) or len(uploaded_names)>0},
        {"item":"Employee written explanation or opportunity to explain","status":has_any(names,["explanation","salaysay","explain","statement"])},
        {"item":"Supporting evidence / uploaded documents","status":len(uploaded_names)>0},
        {"item":"Witness statement if applicable","status":has_any(names,["witness","testimony"])},
        {"item":"System logs / GPS / payroll / CCTV if applicable","status":has_any(names,["gps","log","attendance","payroll","timestamp","cctv","screenshot"])},
        {"item":"Central Command review","status":False},
        {"item":"Notice to Explain before disciplinary decision","status":False},
        {"item":"Final decision only after employee side is considered","status":False},
    ]

    due=int(sum(1 for x in checklist if x["status"])/len(checklist)*100)
    evidence=min(20+len(uploaded_names)*15+(20 if has_any(names,["gps","log","screenshot","payroll"]) else 0),100)
    policy_score=80 if policy[0]["policy_category"]!="No exact company policy match" else 35
    labor_score=75 if labor[0]["labor_category"]!="General labor due process review" else 45
    overall=int((due+evidence+policy_score+labor_score)/4)

    if due < 45:
        next_move="Do not proceed to disciplinary recommendation yet. Request missing due-process documents from coordinator."
    elif has_any(text,["salary","wage","payroll","deduction"]):
        next_move="Treat first as payroll/labor standards verification before worker-directed action."
    elif has_any(text,["theft","fraud","cash","remittance","stolen"]):
        next_move="High-risk allegation. Secure evidence chain and require employee explanation before NTE/decision stage."
    elif overall>=70:
        next_move="Ready for Central Command review. Determine whether NTE is appropriate."
    else:
        next_move="More information required. Request missing documents and allow employee to explain before disciplinary step."

    return {
        "worker":{"name":worker_name,"type":worker_type},
        "submitted_by":submitted_by,
        "area":area,
        "incident_category":category,
        "extracted_content_preview":extracted_text[:2500],
        "policy_assessment":policy,
        "labor_standards_review":labor,
        "due_process_checklist":checklist,
        "scores":{
            "company_policy_match":policy_score,
            "labor_standards_relevance":labor_score,
            "evidence_sufficiency":evidence,
            "due_process_completion":due,
            "overall_readiness":overall
        },
        "missing_requirements":[x["item"] for x in checklist if not x["status"]],
        "recommended_next_move":next_move,
        "safeguard":"No final disciplinary action should be recommended until the worker is given due process under Philippine labor standards."
    }
